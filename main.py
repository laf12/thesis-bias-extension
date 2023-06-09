import cv2
from openpyxl import Workbook
from utils.create_mask import apply_mask
from utils.get_shear_scale import read_shear_scale, read_yaml
from utils.get_shear_value import get_scale_colors, get_average_shear
from utils.get_clamp_distance import detect_clamp
from utils.convert_avi import convert_avi_to_mp4
from utils.convert_to_gif import convert_to_gif
from utils.plot_graph import plot_graph, read_log_file
import numpy as np

# Function to handle trackbar changes
def on_trackbar(position, input_video):
    # Set the video capture to the specified frame position
    input_video.set(cv2.CAP_PROP_POS_FRAMES, position)


def main(video_name, data_loaded):
    # if the video is in AVI format, convert it to MP4
    if video_name.endswith(".avi"):
        convert_avi_to_mp4(video_name, video_name[:-4])
        video_name = video_name[:-4] + ".mp4"

    # read the video
    input_video = cv2.VideoCapture(video_name)

    # get the color scale
    # scale_colors = get_scale_colors(data_loaded['shear_scale_color'])
    scale_colors = get_scale_colors(data_loaded)

    # for each color in the scale, create an image of that color and save it
    i = 0
    for color in scale_colors:
        i += 1
        color_image = np.full((100, 100, 3), color, dtype=np.uint8)
        cv2.imwrite('scale/color_{}.jpg'.format(i), color_image)

    # Get the total number of frames in the video
    total_frames = int(input_video.get(cv2.CAP_PROP_FRAME_COUNT))

    # Create a window to display the video
    cv2.namedWindow('Video')

    fps = int(input_video.get(cv2.CAP_PROP_FPS))
    width = int(input_video.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(input_video.get(cv2.CAP_PROP_FRAME_HEIGHT))

    if data_loaded['video']['save']:
        data_loaded['video']['trackbar'] = False
        # Create a video writer object to save the output video
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        video_file = data_loaded['video']['output_dir'] + data_loaded['video']['output_name']
        output_video = cv2.VideoWriter(video_file, fourcc, fps/16, (width, height))

    # Create a trackbar
    if data_loaded['video']['trackbar']:
        cv2.createTrackbar('Frame', 'Video', 0, total_frames - 1, lambda x: on_trackbar(x, input_video))

    # if log is enabled, create a file to write the log
    if data_loaded['video']['log_values']:
        if data_loaded['video']['excel']:
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            # Set column headers
            ws['A1'] = 'Frame'
            ws['B1'] = 'A'
            ws['C1'] = 'B'
            ws['D1'] = 'C'
            ws['E1'] = 'distance'
        else:
            # create a log file with the same name as the video in the log folder
            log_file = open(data_loaded['video']['log_dir'] + data_loaded['video']['output_name'][:-4] + '.xlsx', 'w')
            log_file.write('Frame A B C distance\n')

    old_distance = 0
    i = 0

    # play the video
    while input_video.isOpened():
        ret, frame = input_video.read()
        
        if not ret:
            break

        frame_copy = frame.copy()

        # Get the current position of the trackbar if it exists
        if data_loaded['video']['trackbar']:
            current_frame = cv2.getTrackbarPos('Frame', 'Video')
            # Set the video capture to the current frame position
            input_video.set(cv2.CAP_PROP_POS_FRAMES, current_frame)

        # Apply the mask to the frame
        red_contrours, frame = apply_mask(frame, 'red', data_loaded)
        green_contrours, frame = apply_mask(frame, 'green', data_loaded)
        blue_contrours, frame = apply_mask(frame, 'blue', data_loaded)

        shear_text_scale = read_shear_scale(frame_copy , data_loaded)
        # print(shear_text_scale)
        red_average_shear = get_average_shear(red_contrours, scale_colors, shear_text_scale, frame_copy)
        green_average_shear = get_average_shear(green_contrours, scale_colors, shear_text_scale, frame_copy)
        blue_average_shear = get_average_shear(blue_contrours, scale_colors, shear_text_scale, frame_copy)

        # print the average shear value
        cv2.putText(frame, 'Red Average Shear: ' + str(red_average_shear), (10, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1, cv2.LINE_AA)
        cv2.putText(frame, 'Green Average Shear: ' + str(green_average_shear), (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1, cv2.LINE_AA)
        cv2.putText(frame, 'Blue Average Shear: ' + str(blue_average_shear), (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 200, 0), 1, cv2.LINE_AA)

        # get the distance between the pins
        # distance = detect_pins(frame_copy, data_loaded)
        distance = detect_clamp(frame_copy, data_loaded)
        cv2.putText(frame, 'Distance: ' + str(distance), (10, 80), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 200, 0), 1, cv2.LINE_AA)
        # cv2.putText(frame, 'Displacement Rate: ' + str((distance - old_distance)/fps), (10, 100), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 200, 0), 1, cv2.LINE_AA)

        # Display the frame
        cv2.imshow('Video', frame)

        # write the log
        if data_loaded['video']['log_values']:
            if data_loaded['video']['excel']:
                row = [i, red_average_shear, green_average_shear, blue_average_shear, distance]
                ws.append(row)
            else:
                log_file.write(str(i) + ' ' + str(red_average_shear) + ' ' + str(green_average_shear) + ' ' + str(blue_average_shear) + ' ' + str(distance) + '\n')

        old_distance = distance
        i += 1

        if data_loaded['video']['save']:
            # Write the frame to the output video
            output_video.write(frame)


        # Exit the loop if 'q' is pressed
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    # release the video capture object
    input_video.release()
    if data_loaded['video']['save']:
        # release the video writer object
        output_video.release()
    # close all windows
    cv2.destroyAllWindows()

    if data_loaded['video']['log_values']:
        if data_loaded['video']['excel']:
            # Save the file
            wb.save(data_loaded['video']['log_dir'] + data_loaded['video']['output_name'][:-4] + '.xlsx')
            wb.close()
        else:
            log_file.close()


if __name__ == '__main__':
    cfg_file = 'cfg/config.yaml'
    data_loaded = read_yaml(cfg_file)
    if data_loaded['video']['get_gif']:
        data_loaded['video']['save'] = True
    file_name = data_loaded['video']['input_dir'] + data_loaded['video']['input_file']
    main(file_name, data_loaded)
    if data_loaded['video']['get_gif']:
        convert_to_gif(data_loaded['video']['output_dir'] + data_loaded['video']['output_name'], data_loaded)
    if data_loaded['video']['plot']:
        log_data = read_log_file(data_loaded['video']['log_dir'] + data_loaded['video']['output_name'][:-4] + '.txt')
        plot_graph(log_data, data_loaded)
